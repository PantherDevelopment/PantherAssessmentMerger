#!/usr/bin/env python3
"""
Panther Assessment Merger
Merges course assessment data into departmental master spreadsheet
"""

import sys
import json
import webbrowser
import urllib.request
import pandas as pd
from pathlib import Path
from PyQt6.QtWidgets import (
    QApplication, QMainWindow, QWidget, QVBoxLayout, QHBoxLayout,
    QPushButton, QLabel, QTextEdit, QFileDialog, QMessageBox,
    QGroupBox, QTreeWidget, QTreeWidgetItem, QTabWidget,
    QTableWidget, QTableWidgetItem, QComboBox, QHeaderView
)
from PyQt6.QtGui import QFont, QColor, QAction
from PyQt6.QtCore import Qt, QThread, pyqtSignal

VERSION = "1.0.0"  # Replaced automatically by GitHub Actions at build time
GITHUB_REPO = "DarbyP/panther-assessment-merger"
SEMESTER_ORDER = {"Fall": 0, "Spring": 1, "Summer": 2}


class UpdateChecker(QThread):
    update_available = pyqtSignal(str, str)

    def run(self):
        try:
            url = f"https://api.github.com/repos/{GITHUB_REPO}/releases/latest"
            with urllib.request.urlopen(url, timeout=5) as response:
                data = json.loads(response.read())
                latest_version = data['tag_name'].lstrip('v')
                if self.is_newer_version(latest_version, VERSION):
                    self.update_available.emit(latest_version, data['html_url'])
        except:
            pass

    def is_newer_version(self, latest, current):
        latest_parts = [int(x) for x in latest.split('.')]
        current_parts = [int(x) for x in current.split('.')]
        return latest_parts > current_parts


# ─────────────────────────────────────────────
#  TAB 1: Merge data into master
# ─────────────────────────────────────────────
class MergeTab(QWidget):
    def __init__(self, log_fn):
        super().__init__()
        self.log = log_fn
        self.master_df = None
        self.master_path = None
        self.master_columns = set()
        self.loaded_files = []
        self.init_ui()

    def init_ui(self):
        layout = QVBoxLayout(self)
        layout.setSpacing(12)

        # Step 1
        master_group = QGroupBox("Step 1: Load Master Spreadsheet")
        master_layout = QHBoxLayout()
        self.load_master_btn = QPushButton("Load Master File")
        self.load_master_btn.clicked.connect(self.load_master_file)
        self.master_label = QLabel("No master file loaded")
        master_layout.addWidget(self.load_master_btn)
        master_layout.addWidget(self.master_label)
        master_layout.addStretch()
        master_group.setLayout(master_layout)
        layout.addWidget(master_group)

        # Step 2
        data_group = QGroupBox("Step 2: Load Your Data File(s)")
        data_layout = QHBoxLayout()
        self.load_data_btn = QPushButton("Add Data File(s)")
        self.load_data_btn.setEnabled(False)
        self.load_data_btn.clicked.connect(self.load_data_files)
        self.clear_data_btn = QPushButton("Clear All Files")
        self.clear_data_btn.setEnabled(False)
        self.clear_data_btn.clicked.connect(self.clear_data_files)
        data_layout.addWidget(self.load_data_btn)
        data_layout.addWidget(self.clear_data_btn)
        data_layout.addStretch()
        data_group.setLayout(data_layout)
        layout.addWidget(data_group)

        # Step 3: Preview
        preview_group = QGroupBox("Step 3: Column Match Preview")
        preview_layout = QVBoxLayout()
        self.preview_tree = QTreeWidget()
        self.preview_tree.setHeaderLabels(["File / Column", "Status"])
        self.preview_tree.setColumnWidth(0, 550)
        self.preview_tree.setMinimumHeight(220)
        self.preview_tree.setAlternatingRowColors(True)
        preview_layout.addWidget(self.preview_tree)
        preview_group.setLayout(preview_layout)
        layout.addWidget(preview_group)

        # Step 4
        action_group = QGroupBox("Step 4: Merge")
        action_layout = QHBoxLayout()
        self.merge_btn = QPushButton("Merge & Save")
        self.merge_btn.setEnabled(False)
        self.merge_btn.clicked.connect(self.merge_and_save)
        action_layout.addWidget(self.merge_btn)
        action_layout.addStretch()
        action_group.setLayout(action_layout)
        layout.addWidget(action_group)

    def load_master_file(self):
        file_path, _ = QFileDialog.getOpenFileName(
            self, "Select Master Spreadsheet", "", "Excel Files (*.xlsx *.xls)"
        )
        if not file_path:
            return
        try:
            self.master_df = pd.read_excel(file_path)
            self.master_path = Path(file_path)
            if 'STUDENT_ID' not in self.master_df.columns:
                QMessageBox.warning(self, "Missing Column", "Master file must have a STUDENT_ID column")
                self.master_df = None
                return
            self.master_columns = set(self.master_df.columns) - {'STUDENT_ID'}
            self.master_label.setText(f"Loaded: {self.master_path.name}")
            self.log(f"[Merge] Master file loaded: {self.master_path.name} ({len(self.master_df)} students, {len(self.master_columns)} assessment columns)")
            self.load_data_btn.setEnabled(True)
            if self.loaded_files:
                self.analyze_and_preview()
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Failed to load master file:\n{str(e)}")
            self.log(f"[Merge] ERROR loading master file: {str(e)}")

    def load_data_files(self):
        file_paths, _ = QFileDialog.getOpenFileNames(
            self, "Select Data File(s)", "", "Excel Files (*.xlsx *.xls);;CSV Files (*.csv)"
        )
        if not file_paths:
            return
        for file_path in file_paths:
            try:
                df = pd.read_csv(file_path) if file_path.endswith('.csv') else pd.read_excel(file_path)
                if 'STUDENT_ID' not in df.columns:
                    QMessageBox.warning(self, "Missing Column",
                        f"{Path(file_path).name} has no STUDENT_ID column — skipped")
                    continue
                self.loaded_files.append({'path': Path(file_path), 'df': df})
                self.log(f"[Merge] Loaded: {Path(file_path).name} ({len(df)} students)")
            except Exception as e:
                self.log(f"[Merge] ERROR loading {Path(file_path).name}: {str(e)}")
        if self.loaded_files:
            self.clear_data_btn.setEnabled(True)
            self.analyze_and_preview()

    def analyze_and_preview(self):
        self.preview_tree.clear()
        any_matches = False
        for file_info in self.loaded_files:
            file_cols = set(file_info['df'].columns) - {'STUDENT_ID'}
            matched = sorted(file_cols & self.master_columns)
            unmatched = sorted(file_cols - self.master_columns)
            file_info['matched_cols'] = matched
            file_info['unmatched_cols'] = unmatched

            file_item = QTreeWidgetItem(self.preview_tree)
            file_item.setText(0, file_info['path'].name)
            if not matched:
                file_item.setText(1, "⚠️  No matching columns")
                file_item.setForeground(1, QColor("#c0392b"))
                file_item.setForeground(0, QColor("#c0392b"))
            elif unmatched:
                file_item.setText(1, f"✓  {len(matched)} matched, {len(unmatched)} unmatched (ignored)")
                file_item.setForeground(1, QColor("#d35400"))
                any_matches = True
            else:
                file_item.setText(1, f"✓  {len(matched)} columns matched")
                file_item.setForeground(1, QColor("#27ae60"))
                any_matches = True

            if matched:
                mp = QTreeWidgetItem(file_item)
                mp.setText(0, "  Columns to merge:")
                mp.setForeground(0, QColor("#27ae60"))
                for col in matched:
                    c = QTreeWidgetItem(mp)
                    c.setText(0, f"      {col}")
                    c.setForeground(0, QColor("#27ae60"))
            if unmatched:
                up = QTreeWidgetItem(file_item)
                up.setText(0, "  Columns not in master (ignored):")
                up.setForeground(0, QColor("#c0392b"))
                for col in unmatched:
                    c = QTreeWidgetItem(up)
                    c.setText(0, f"      {col}")
                    c.setForeground(0, QColor("#c0392b"))
                self.log(f"[Merge] WARNING: {file_info['path'].name} has unmatched columns ignored: {', '.join(unmatched)}")

            file_item.setExpanded(True)
            for i in range(file_item.childCount()):
                file_item.child(i).setExpanded(True)

        self.merge_btn.setEnabled(any_matches)
        if not any_matches:
            self.log("[Merge] No matching columns found across any loaded files.")

    def clear_data_files(self):
        self.loaded_files = []
        self.preview_tree.clear()
        self.merge_btn.setEnabled(False)
        self.clear_data_btn.setEnabled(False)
        self.log("[Merge] Cleared all data files.")

    def merge_and_save(self):
        try:
            from openpyxl import load_workbook
            wb = load_workbook(self.master_path)
            ws = wb.active
            headers = {cell.value: col_idx for col_idx, cell in enumerate(ws[1], start=1) if cell.value}
            if 'STUDENT_ID' not in headers:
                QMessageBox.warning(self, "Error", "STUDENT_ID column not found in master file")
                return
            student_id_col = headers['STUDENT_ID']
            student_row_map = {
                ws.cell(r, student_id_col).value: r
                for r in range(2, ws.max_row + 1)
                if ws.cell(r, student_id_col).value is not None
            }
            total_updated = set()
            all_missing = set()
            for file_info in self.loaded_files:
                if not file_info['matched_cols']:
                    continue
                df = file_info['df'].drop_duplicates(subset=['STUDENT_ID'], keep='last')
                for _, user_row in df.iterrows():
                    sid = user_row['STUDENT_ID']
                    found_row = student_row_map.get(sid)
                    if found_row:
                        for col in file_info['matched_cols']:
                            if pd.notna(user_row.get(col)):
                                ws.cell(found_row, headers[col]).value = user_row[col]
                        total_updated.add(sid)
                    else:
                        all_missing.add(sid)
            wb.save(self.master_path)
            report = f"\n{'='*50}\nMERGE COMPLETE\n{'='*50}\n"
            report += f"Students updated: {len(total_updated)}\n"
            report += f"Students NOT in master: {len(all_missing)}\n"
            if all_missing:
                report += "\nMissing Student IDs:\n" + "".join(f"  • {s}\n" for s in sorted(all_missing))
            report += f"\nMaster file updated (formatting preserved):\n{self.master_path}"
            self.log(report)
            QMessageBox.information(self, "Merge Complete",
                f"Successfully updated {len(total_updated)} students!\n\n"
                f"{len(all_missing)} students not found in master.\n\n"
                "Master file updated with formatting preserved.")
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Failed to merge data:\n{str(e)}")
            self.log(f"[Merge] ERROR during merge: {str(e)}")


# ─────────────────────────────────────────────
#  TAB 2: Combine semester files into yearly
# ─────────────────────────────────────────────
class YearlyTab(QWidget):
    def __init__(self, log_fn):
        super().__init__()
        self.log = log_fn
        self.semester_files = []
        self.init_ui()

    def init_ui(self):
        layout = QVBoxLayout(self)
        layout.setSpacing(12)

        # Step 1
        add_group = QGroupBox("Step 1: Load Files to Combine")
        add_layout = QVBoxLayout()
        info_label = QLabel(
            "Load the semester master files you want to combine. "
            "Use 'Already Combined' for a file that contains combined data from Fall and Spring. "
            "Only use this option to add Summer to a combined Fall/Spring dataset."
        )
        info_label.setWordWrap(True)
        info_label.setStyleSheet("color: #555; padding: 4px;")
        add_layout.addWidget(info_label)
        btn_layout = QHBoxLayout()
        self.add_btn = QPushButton("Add File(s)")
        self.add_btn.clicked.connect(self.add_semester_files)
        self.clear_btn = QPushButton("Clear All")
        self.clear_btn.setEnabled(False)
        self.clear_btn.clicked.connect(self.clear_files)
        btn_layout.addWidget(self.add_btn)
        btn_layout.addWidget(self.clear_btn)
        btn_layout.addStretch()
        add_layout.addLayout(btn_layout)
        add_group.setLayout(add_layout)
        layout.addWidget(add_group)

        # Step 2
        assign_group = QGroupBox("Step 2: Assign Type to Each File")
        assign_layout = QVBoxLayout()
        assign_layout.addWidget(QLabel(
            "Semester priority order (lowest to highest): Fall → Spring → Summer → Already Combined.\n"
            "Only one 'Already Combined' file is allowed. It will be treated as earlier data, "
            "overwritten by any single-semester file."
        ))
        self.file_table = QTableWidget(0, 3)
        self.file_table.setHorizontalHeaderLabels(["File", "Students", "Type"])
        self.file_table.horizontalHeader().setSectionResizeMode(0, QHeaderView.ResizeMode.Stretch)
        self.file_table.horizontalHeader().setSectionResizeMode(1, QHeaderView.ResizeMode.ResizeToContents)
        self.file_table.horizontalHeader().setSectionResizeMode(2, QHeaderView.ResizeMode.ResizeToContents)
        self.file_table.setMinimumHeight(160)
        self.file_table.setEditTriggers(QTableWidget.EditTrigger.NoEditTriggers)
        assign_layout.addWidget(self.file_table)
        assign_group.setLayout(assign_layout)
        layout.addWidget(assign_group)

        # Step 3
        output_group = QGroupBox("Step 3: Save Yearly File")
        output_layout = QHBoxLayout()
        self.combine_btn = QPushButton("Combine & Save Yearly File")
        self.combine_btn.setEnabled(False)
        self.combine_btn.clicked.connect(self.combine_and_save)
        output_layout.addWidget(self.combine_btn)
        output_layout.addStretch()
        output_group.setLayout(output_layout)
        layout.addWidget(output_group)

    def add_semester_files(self):
        file_paths, _ = QFileDialog.getOpenFileNames(
            self, "Select File(s)", "", "Excel Files (*.xlsx *.xls)"
        )
        if not file_paths:
            return
        for file_path in file_paths:
            try:
                df = pd.read_excel(file_path)
                if 'STUDENT_ID' not in df.columns:
                    QMessageBox.warning(self, "Missing Column",
                        f"{Path(file_path).name} has no STUDENT_ID column — skipped")
                    continue
                self.semester_files.append({'path': Path(file_path), 'df': df, 'semester': 'Fall'})
                self.log(f"[Yearly] Loaded: {Path(file_path).name} ({len(df)} students)")
            except Exception as e:
                self.log(f"[Yearly] ERROR loading {Path(file_path).name}: {str(e)}")
        self.refresh_table()
        self.clear_btn.setEnabled(True)
        self.combine_btn.setEnabled(len(self.semester_files) >= 2)

    def refresh_table(self):
        self.file_table.setRowCount(0)
        for i, file_info in enumerate(self.semester_files):
            self.file_table.insertRow(i)
            self.file_table.setItem(i, 0, QTableWidgetItem(file_info['path'].name))
            self.file_table.setItem(i, 1, QTableWidgetItem(str(len(file_info['df']))))
            combo = QComboBox()
            combo.addItems(["Fall", "Spring", "Summer", "Already Combined"])
            combo.setCurrentText(file_info['semester'])
            combo.currentTextChanged.connect(lambda val, idx=i: self.update_semester(idx, val))
            self.file_table.setCellWidget(i, 2, combo)

    def update_semester(self, idx, semester):
        if idx < len(self.semester_files):
            self.semester_files[idx]['semester'] = semester

    def clear_files(self):
        self.semester_files = []
        self.file_table.setRowCount(0)
        self.combine_btn.setEnabled(False)
        self.clear_btn.setEnabled(False)
        self.log("[Yearly] Cleared all semester files.")

    def combine_and_save(self):
        try:
            # Validate: only one "Already Combined" allowed
            combined_count = sum(1 for f in self.semester_files if f['semester'] == 'Already Combined')
            if combined_count > 1:
                QMessageBox.warning(self, "Invalid Selection",
                    "Only one 'Already Combined' file can be used at a time.\n\n"
                    "Please tag only one file as 'Already Combined' and set the other(s) "
                    "to their specific semester (Fall, Spring, or Summer).")
                return

            # Priority order: Fall=0, Spring=1, Summer=2, Already Combined=3 (highest, processed last)
            order = {"Fall": 0, "Spring": 1, "Summer": 2, "Already Combined": 3}
            sorted_files = sorted(self.semester_files, key=lambda x: order.get(x['semester'], 99))

            self.log(f"[Yearly] Processing order: {chr(32).join(f['path'].name + ' (' + f['semester'] + ')' for f in sorted_files)}")

            # Union of all columns
            all_cols = set()
            for f in sorted_files:
                all_cols.update(f['df'].columns)
            all_cols.discard('STUDENT_ID')
            all_cols = sorted(all_cols)

            # Build combined dict — later files overwrite earlier ones column by column
            combined = {}
            for file_info in sorted_files:
                df = file_info['df'].drop_duplicates(subset=['STUDENT_ID'], keep='last')
                for _, row in df.iterrows():
                    sid = row['STUDENT_ID']
                    if sid not in combined:
                        combined[sid] = {}
                    for col in all_cols:
                        if col in row and pd.notna(row[col]):
                            combined[sid][col] = row[col]

            rows = []
            for sid, col_data in combined.items():
                row = {'STUDENT_ID': sid}
                row.update(col_data)
                rows.append(row)

            output_df = pd.DataFrame(rows, columns=['STUDENT_ID'] + all_cols)
            output_df = output_df.sort_values('STUDENT_ID').reset_index(drop=True)

            save_path, _ = QFileDialog.getSaveFileName(
                self, "Save Yearly File", "yearly_master.xlsx", "Excel Files (*.xlsx)"
            )
            if not save_path:
                return

            output_df.to_excel(save_path, index=False)

            report = f"\n{'='*50}\nYEARLY COMBINE COMPLETE\n{'='*50}\n"
            report += f"Total unique students: {len(output_df)}\n"
            report += f"Files combined: {', '.join(f['path'].name + ' (' + f['semester'] + ')' for f in sorted_files)}\n"
            report += f"Saved to: {save_path}"
            self.log(report)

            QMessageBox.information(self, "Done",
                f"Yearly file created!\n\n"
                f"{len(output_df)} unique students across {len(sorted_files)} files.\n\n"
                f"Saved to: {Path(save_path).name}")

        except Exception as e:
            QMessageBox.critical(self, "Error", f"Failed to combine files:\n{str(e)}")
            self.log(f"[Yearly] ERROR: {str(e)}")



USER_GUIDE = """
<b>Tab 1: Merge Data into Master</b><br><br>

Use this tab to add faculty assessment data into the departmental master spreadsheet.<br><br>

<b>Step 1 — Load Master File</b><br>
Click "Load Master File" and select the departmental master spreadsheet (.xlsx). 
The file must contain a STUDENT_ID column.<br><br>

<b>Step 2 — Load Your Data File(s)</b><br>
Click "Add Data File(s)" and select one or more faculty data files. 
You can select multiple files at once using Cmd+Click. 
Each file must contain a STUDENT_ID column and columns matching the master spreadsheet.<br><br>

<b>Step 3 — Review Column Match Preview</b><br>
The preview shows which columns in each file match the master spreadsheet:<br>
• <b>Green</b> — columns that will be merged<br>
• <b>Orange</b> — some columns match, some do not (unmatched columns are ignored)<br>
• <b>Red</b> — no matching columns found (file will be skipped)<br><br>

<b>Step 4 — Merge & Save</b><br>
Click "Merge & Save" to update the master file. 
Existing data is overwritten only where your files have values — blank cells in your data 
will never overwrite existing data in the master. 
The master file's formatting, colors, and structure are fully preserved.<br><br>

After merging, the Status Log will show how many students were updated and list 
any student IDs from your data that were not found in the master spreadsheet.<br><br>

<hr><br>

<b>Tab 2: Combine Semesters into Yearly</b><br><br>

Use this tab to combine semester master files into a single yearly file. 
All students from all semester files will be included in the output.<br><br>

<b>Step 1 — Load Files</b><br>
Click "Add File(s)" and select the semester master files to combine.<br><br>

<b>Step 2 — Assign Type to Each File</b><br>
Use the dropdown next to each file to indicate which semester it represents: 
Fall, Spring, Summer, or Already Combined.<br><br>

Use <b>"Already Combined"</b> only when a file already contains merged data from Fall and Spring 
and you need to add Summer data to it. Only one "Already Combined" file is allowed per operation.<br><br>

Priority order (lowest to highest): Fall → Spring → Summer → Already Combined. 
For students appearing in multiple files, data from the higher-priority file will be used 
(column by column — blank cells never overwrite existing data).<br><br>

<b>Step 3 — Save Yearly File</b><br>
Click "Combine & Save Yearly File" and choose where to save the output. 
The original semester files are not modified.<br><br>

<hr><br>

<b>Tips</b><br>
• Always keep a backup of the master file before merging.<br>
• Student IDs must match exactly between files (same format, no leading/trailing spaces).<br>
• For questions or issues, contact Darby Proctor, Ph.D.
"""

# ─────────────────────────────────────────────
#  Main Window
# ─────────────────────────────────────────────
class AssessmentMerger(QMainWindow):
    def __init__(self):
        super().__init__()
        self.init_ui()

    def init_ui(self):
        self.setWindowTitle("Panther Assessment Merger")
        self.setGeometry(100, 100, 950, 850)

        # Menu bar
        menubar = self.menuBar()

        help_menu = menubar.addMenu("Help")

        user_guide_action = QAction("User Guide", self)
        user_guide_action.triggered.connect(self.show_user_guide)
        help_menu.addAction(user_guide_action)

        about_action = QAction("About Panther Assessment Merger", self)
        about_action.triggered.connect(self.show_about)
        help_menu.addAction(about_action)

        check_updates_action = QAction("Check for Updates", self)
        check_updates_action.triggered.connect(self.check_updates_manual)
        help_menu.addAction(check_updates_action)

        central_widget = QWidget()
        self.setCentralWidget(central_widget)
        main_layout = QVBoxLayout(central_widget)
        main_layout.setSpacing(10)

        # Title
        title = QLabel("Panther Assessment Merger")
        title_font = QFont()
        title_font.setPointSize(16)
        title_font.setBold(True)
        title.setFont(title_font)
        title.setAlignment(Qt.AlignmentFlag.AlignCenter)
        main_layout.addWidget(title)

        # Tabs
        self.tabs = QTabWidget()
        self.merge_tab = MergeTab(self.log)
        self.yearly_tab = YearlyTab(self.log)
        self.tabs.addTab(self.merge_tab, "Merge Data into Master")
        self.tabs.addTab(self.yearly_tab, "Combine Semesters into Yearly")
        main_layout.addWidget(self.tabs)

        # Shared status log
        log_group = QGroupBox("Status Log")
        log_layout = QVBoxLayout()
        self.log_text = QTextEdit()
        self.log_text.setReadOnly(True)
        self.log_text.setMinimumHeight(180)
        log_layout.addWidget(self.log_text)
        log_group.setLayout(log_layout)
        main_layout.addWidget(log_group)

        # Credits
        credit_label = QLabel("Developed by Darby Proctor, Ph.D.")
        credit_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        credit_label.setStyleSheet("color: gray; font-size: 10px; padding: 4px;")
        main_layout.addWidget(credit_label)

        self.version_label = QLabel(f"Version {VERSION}")
        self.version_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self.version_label.setStyleSheet("color: gray; font-size: 9px;")
        main_layout.addWidget(self.version_label)

        self.log("Ready.")

        self.update_checker = UpdateChecker()
        self.update_checker.update_available.connect(self.show_update_notification)
        self.update_checker.start()

    def log(self, message):
        self.log_text.append(message)

    def show_about(self):
        QMessageBox.about(self, "About Panther Assessment Merger",
            f"<b>Panther Assessment Merger</b><br>"
            f"Version {VERSION}<br><br>"
            f"Merges faculty course assessment data into the departmental "
            f"master spreadsheet and combines semester files into yearly reports.<br><br>"
            f"Developed by Darby Proctor, Ph.D.<br>"
            f"Florida Institute of Technology<br><br>"
            f"<a href='https://github.com/{GITHUB_REPO}/releases'>Check for updates</a>"
        )

    def show_user_guide(self):
        guide = QMessageBox(self)
        guide.setWindowTitle("User Guide")
        guide.setIcon(QMessageBox.Icon.NoIcon)
        guide.setText("<b>Panther Assessment Merger — User Guide</b>")
        guide.setInformativeText(USER_GUIDE)
        guide.setStandardButtons(QMessageBox.StandardButton.Ok)
        guide.exec()

    def check_updates_manual(self):
        try:
            url = f"https://api.github.com/repos/{GITHUB_REPO}/releases/latest"
            with urllib.request.urlopen(url, timeout=5) as response:
                data = json.loads(response.read())
                latest_version = data['tag_name'].lstrip('v')
                latest_parts = [int(x) for x in latest_version.split('.')]
                current_parts = [int(x) for x in VERSION.split('.')]
                if latest_parts > current_parts:
                    msg = QMessageBox(self)
                    msg.setWindowTitle("Update Available")
                    msg.setText(f"A new version (v{latest_version}) is available!")
                    msg.setInformativeText("Would you like to download it?")
                    msg.setStandardButtons(QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No)
                    if msg.exec() == QMessageBox.StandardButton.Yes:
                        webbrowser.open(data['html_url'])
                else:
                    QMessageBox.information(self, "Up to Date",
                        f"You are running the latest version (v{VERSION}).")
        except Exception:
            QMessageBox.warning(self, "Check Failed",
                "Could not check for updates. Please check your internet connection.")

    def show_update_notification(self, new_version, download_url):
        self.version_label.setText(f"Version {VERSION} — Update available: v{new_version}")
        self.version_label.setStyleSheet("color: #d35400; font-size: 9px; font-weight: bold;")
        msg = QMessageBox(self)
        msg.setIcon(QMessageBox.Icon.Information)
        msg.setWindowTitle("Update Available")
        msg.setText(f"A new version (v{new_version}) is available!")
        msg.setInformativeText("Would you like to download it?")
        msg.setStandardButtons(QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No)
        if msg.exec() == QMessageBox.StandardButton.Yes:
            webbrowser.open(download_url)


def main():
    app = QApplication(sys.argv)
    window = AssessmentMerger()
    window.show()
    sys.exit(app.exec())


if __name__ == "__main__":
    main()